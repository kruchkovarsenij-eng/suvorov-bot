import sys
import os
import subprocess
import asyncio
import logging
from datetime import datetime

# ---------- Ранний вывод и автоустановка ----------
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)
print("--- [DOCKER START] ИНИЦИАЛИЗАЦИЯ СИСТЕМНОГО ОКРУЖЕНИЯ ---", flush=True)

try:
    import aiogram
    import openpyxl
    import aiohttp
except ModuleNotFoundError:
    print("Кэш сборщика пуст. Принудительно устанавливаю библиотеки...", flush=True)
    subprocess.check_call([sys.executable, "-m", "pip", "install", "aiogram>=3.0.0", "openpyxl", "aiohttp"])
    print("Установка успешно завершена! ✅", flush=True)

from aiogram import Bot, Dispatcher, F, Router
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.filters import CommandStart, Command
from aiogram.client.default import DefaultBotProperties
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.webhook.aiohttp_server import SimpleRequestHandler
from aiohttp import web

print("--- [DOCKER START] ЗАПУСК AIOGRAM ЯДРА ---", flush=True)

# ---------- Конфигурация ----------
TOKEN = "8959504034:AAFTvRop6ApDFX6dCnngx50LmEye_WtZ6C4"
ADMIN_CHAT_ID = 8743677274

DATA_DIR = "data"
EXCEL_FILE = os.path.join(DATA_DIR, "diagnostics_results.xlsx")

WEBHOOK_HOST = os.getenv("WEBHOOK_HOST", "https://suvorov-bot-production.up.railway.app")
WEBHOOK_PATH = "/webhook"
WEBHOOK_URL = f"{WEBHOOK_HOST}{WEBHOOK_PATH}"
WEBHOOK_PORT = int(os.getenv("PORT", 8080))
USE_POLLING = os.getenv("USE_POLLING", "false").lower() == "true"

# ---------- Логирование ----------
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger('aiohttp.access').setLevel(logging.DEBUG)

# ---------- Бот и диспатчер ----------
session = AiohttpSession()
bot = Bot(token=TOKEN, session=session, default=DefaultBotProperties(parse_mode="Markdown"))
dp = Dispatcher()
router = Router()

# ---------- Вопросы (полный список) ----------
QUESTIONS = [
    {"s": "Личные данные", "q": "Как вас зовут? (ФИО)", "t": "open"},
    {"s": "Личные данные", "q": "Сколько лет вы руководите этой компанией?", "t": "open"},
    {"s": "Стратегия и видение", "q": "Где вы видите компанию через 3–5 лет?\n\n 💡 Конкретно: выручка, доля рынка, структура — не «расти и развиваться»", "t": "open"},
    {"s": "Стратегия и видение", "q": "Какие 3 стратегических приоритета вы бы поставили на первое место прямо сейчас?", "t": "open"},
    {"s": "Стратегия и видение", "q": "Что является главным драйвером роста компании?\n Что — якорем, который тормозит?", "t": "open"},
    {"s": "Стратегия и видение", "q": "Какие возможности вы видите на рынке, которые используете недостаточно?", "t": "open"},
    {"s": "Стратегия и видение", "q": "Что изменится в компании через 2 года, если ничего не менять прямо сейчас?", "t": "open"},
    {"s": "Состояние компании", "q": "Как вы оцениваете текущее состояние компании?\n\n 1 — кризис, 10 — отличная форма", "t": "scale"},
    {"s": "Состояние компании", "q": "Что у нас получается лучше всего как у команды?", "t": "open"},
    {"s": "Состояние компании", "q": "В каких областях мы чаще всего сталкиваемся с трудностями?", "t": "open"},
    {"s": "Состояние компании", "q": "Что является самым большим препятствием для развития компании?", "t": "open"},
    {"s": "Состояние компании", "q": "Какие ключевые угрозы вы видите для бизнеса на горизонте 1–3 лет?", "t": "open"},
    {"s": "Состояние компании", "q": "Какие возможности мы упускаем или используем недостаточно?", "t": "open"},
    {"s": "Команда", "q": "Насколько высок уровень доверия в управленческой команде?\n\n 1 — нет доверия, 10 — полное доверие", "t": "scale"},
    {"s": "Команда", "q": "Насколько открыто мы обсуждаем проблемы и конфликты?\n\n 1 — избегаем, 10 — обсуждаем открыто", "t": "scale"},
    {"s": "Команда", "q": "Оцените взаимодействие между подразделениями\n\n 1 — нет взаимодействия, 10 — отличное", "t": "scale"},
    {"s": "Команда", "q": "Что укрепляет команду, а что её ослабляет?", "t": "open"},
    {"s": "Команда", "q": "Какие сильные стороны нашей команды стоит развивать?", "t": "open"},
    {"s": "Команда", "q": "Какие слабые стороны мешают работать эффективнее?", "t": "open"},
    {"s": "Операционка", "q": "Как в компании контролируется выполнение задач?\n\n 💡 Что реально происходит, когда задача не выполнена в срок?", "t": "open"},
    {"s": "Операционка", "q": "Как часто задачи выполняются без напоминаний?\n\n 1 — никогда, 10 — всегда", "t": "scale"},
    {"s": "Операционка", "q": "Всегда ли понятно, кто за что отвечает?\n\n 💡 Есть ли случаи, когда виноватых нет?", "t": "open"},
    {"s": "Операционка", "q": "Бывает ли: ответственный назначен — результата нет — последствий тоже нет?\n\n 💡 Опишите конкретный пример", "t": "open"},
    {"s": "Операционка", "q": "Назовите цели компании на текущий год / квартал\n\n 💡 Без подготовки — то, что знаете прямо сейчас", "t": "open"},
    {"s": "Операционка", "q": "Насколько планы соответствуют реальности выполнения?\n\n 1 — планы не выполняются, 10 — всегда в срок", "t": "scale"},
    {"s": "Операционка", "q": "Насколько совещания в компании результативны?\n\n 1 — пустая трата времени, 10 — максимально результативны", "t": "scale"},
    {"s": "Операционка", "q": "После совещаний фиксируются ли решения и ответственные? Как это работает на практике?", "t": "open"},
    {"s": "Система управления", "q": "Насколько компания управляема без вашего личного участия?\n\n 1 — без меня всё остановится, 10 — работает самостоятельно", "t": "scale"},
    {"s": "Система управления", "q": "Где вы лично являетесь узким местом системы управления?\n\n 💡 В чём вы сами тормозите компанию — честно", "t": "open"},
    {"s": "Система управления", "q": "Какие решения вы вынуждены принимать сами, хотя могли бы делегировать? Почему не делегируете?", "t": "open"},
    {"s": "Система управления", "q": "Что в компании держится только на вас — и почему это опасно?", "t": "open"},
    {"s": "Система управления", "q": "Какое управленческое решение вы откладываете уже давно, хотя знаете, что его нужно принять?", "t": "open"},
    {"s": "Система управления", "q": "Есть ли в команде люди, которые тормозят систему? Что с этим делается?", "t": "open"},
    {"s": "Управленческая команда", "q": "Как вы оцениваете качество своей управленческой команды в целом?\n\n 1 — команда слабая, 10 — команда сильная", "t": "scale"},
    {"s": "Управленческая команда", "q": "Кто из команды точно на своём месте? Кто — нет? Почему до сих пор не изменили ситуацию?", "t": "open"},
    {"s": "Управленческая команда", "q": "Кого из команды вы бы взяли с собой, если бы начинали всё заново? Почему?", "t": "open"},
    {"s": "Управленческая команда", "q": "Что происходит в компании, когда вас нет? Приведите конкретный пример.", "t": "open"},
    {"s": "Деньги и потери", "q": "Где company теряет деньги прямо сейчас, но причина ещё не устранена?", "t": "open"},
    {"s": "Деньги и потери", "q": "Что является самым узким горлышком в компании прямо сейчас?", "t": "open"},
    {"s": "Самооценка лидера", "q": "Какой управленческий стиль вы используете чаще всего?", "t": "choice", "opts": ["Директивный (я решаю, команда выполняет)", "Делегирующий (ставлю задачу, доверяю результат)", "Коучинговый (развиваю людей через вопросы)", "Хаотичный (по ситуации, системы нет)"]},
    {"s": "Самооценка лидера", "q": "Что вы готовы изменить в собственном стиле управления?\n\n 💡 Конкретно — не «стать лучше», а что именно и в какой срок", "t": "open"},
    {"s": "Самооценка лидера", "q": "Назовите одну вещь, которую вы бы изменили в компании завтра, если бы не было сопротивления.", "t": "open"},
    {"s": "Самооценка лидера", "q": "Что вы хотите получить от диагностики и консалтинга?\n\n 💡 Конкретная метрика успеха, которую готовы зафиксировать как результат", "t": "open"},
]

class Survey(StatesGroup):
    answering = State()

user_sessions = {}

# ---------- Excel ----------
def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ответы"
    hf = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hf_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers = ["Дата прохождения", "ID Пользователя", "Никнейм (username)"]
    for i, q in enumerate(QUESTIONS):
        headers.append(f"Вопрос {i+1}: {q['q'].replace('\n', ' ')} [{q['s']}]")
    ws.append(headers)
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = hf, hf_fill, ca
    ws.row_dimensions[1].height = 35
    wb.save(EXCEL_FILE)

def save_to_excel_final(user_id, username, answers_list):
    init_excel()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        row_data = [datetime.now().strftime("%d.%m.%Y %H:%M"), str(user_id), f"@{username}" if username else "—"]
        for i in range(len(QUESTIONS)):
            row_data.append(str(answers_list[i]) if i < len(answers_list) else "—")
        ws.append(row_data)
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 60)
        wb.save(EXCEL_FILE)
        logging.info(f"✅ Ответы пользователя {user_id} записаны в Excel")
    except Exception as e:
        logging.error(f"Excel error: {e}")

# ---------- Клавиатуры ----------
def get_scale_keyboard():
    buttons1 = [InlineKeyboardButton(text=str(i), callback_data=f"scale_{i}") for i in range(1, 6)]
    buttons2 = [InlineKeyboardButton(text=str(i), callback_data=f"scale_{i}") for i in range(6, 11)]
    return InlineKeyboardMarkup(inline_keyboard=[buttons1, buttons2])

def get_choice_keyboard(opts):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=opt, callback_data=f"choice_{i}")] for i, opt in enumerate(opts)
    ])

def format_report(user_id, username, answers_list):
    lines = [
        "📋 ДИАГНОСТИКА — «Пластик Руси»",
        f"👤 ID: {user_id}",
        f"🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        ""
    ]
    current_section = ""
    for i, q in enumerate(QUESTIONS):
        if q["s"] != current_section:
            current_section = q["s"]
            lines.extend([f"\n{'━'*28}", f"📌 {current_section.upper()}", f"{'━'*28}"])
        ans = answers_list[i] if i < len(answers_list) else "—"
        clean_q = q['q'].replace('\n\n', ' ').replace('\n', ' ')
        lines.append(f"❓ {clean_q}")
        lines.append(f"➜ {ans}")
    return "\n".join(lines)

def get_progress_bar(current_question, total_questions):
    bar_length = 20
    filled = int(bar_length * current_question / total_questions)
    bar = '▓' * filled + '░' * (bar_length - filled)
    return f"Прогресс: {bar} [{current_question}/{total_questions}]"

# ---------- Хендлеры ----------
@router.message(CommandStart())
async def start(message: Message, state: FSMContext):
    welcome = (
        "👋 *Привет! Это диагностика управленческой команды «Пластик Руси»*\n\n"
        "📋 *Что будет:*\n"
        f"• Вам предстоит ответить на *{len(QUESTIONS)} вопросов* по ключевым блокам управления.\n"
        "• Вопросы разного типа: открытые, оценка по шкале 1–10, выбор варианта.\n"
        "• В каждом сообщении будет *прогресс‑бар*, чтобы вы видели, сколько пройдено.\n\n"
        "⏱ *Сколько времени займёт:* примерно 20–30 минут (зависит от темпа ответов).\n"
        "🧠 *Совет:* отвечайте честно и без подготовки — важна реальная картина, а не «правильные» ответы.\n\n"
        "📊 *Результат:* после завершения вы получите полный текстовый отчёт, "
        "а все данные сохранятся в единую таблицу для анализа.\n\n"
        "👉 Чтобы начать, просто продолжайте — первый вопрос уже здесь:"
    )
    await message.answer(welcome, parse_mode="Markdown")

    await state.set_state(Survey.answering)
    user_sessions[message.from_user.id] = {
        "current_q": 0,
        "answers": []
    }

    q = QUESTIONS[0]
    progress = get_progress_bar(1, len(QUESTIONS))
    text = f"{progress}\n\n{q['q']}"
    if q["t"] == "scale":
        await message.answer(text, reply_markup=get_scale_keyboard())
    elif q["t"] == "choice":
        await message.answer(text, reply_markup=get_choice_keyboard(q["opts"]))
    else:
        await message.answer(text)

@router.callback_query(F.data.startswith("scale_"))
async def scale_answer(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    session = user_sessions.get(user_id)
    if not session:
        await callback.answer("Начните с /start")
        return
    value = callback.data.split("_")[1]
    session["answers"].append(value)
    session["current_q"] += 1
    await next_question(callback.message, user_id, state)
    await callback.answer()

@router.callback_query(F.data.startswith("choice_"))
async def choice_answer(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    session = user_sessions.get(user_id)
    if not session:
        await callback.answer("Начните с /start")
        return
    idx = int(callback.data.split("_")[1])
    q_idx = session["current_q"]
    answer_text = QUESTIONS[q_idx]["opts"][idx]
    session["answers"].append(answer_text)
    session["current_q"] += 1
    await next_question(callback.message, user_id, state)
    await callback.answer()

@router.message(Survey.answering)
async def open_answer(message: Message, state: FSMContext):
    user_id = message.from_user.id
    session = user_sessions.get(user_id)
    if not session:
        await message.answer("Начните с /start")
        return
    session["answers"].append(message.text)
    session["current_q"] += 1
    await next_question(message, user_id, state)

async def next_question(msg, user_id, state: FSMContext):
    session = user_sessions[user_id]
    idx = session["current_q"]
    if idx >= len(QUESTIONS):
        save_to_excel_final(user_id, msg.from_user.username, session["answers"])
        report = format_report(user_id, msg.from_user.username, session["answers"])

        # Разбиваем длинный отчёт на части по 4000 символов (лимит Telegram 4096)
        max_len = 4000
        parts = [report[i:i+max_len] for i in range(0, len(report), max_len)]

        await msg.answer("✅ Спасибо за ответы! Вот ваша диагностика:")
        for part in parts:
            await msg.answer(part)

        await state.clear()
        return

    q = QUESTIONS[idx]
    progress = get_progress_bar(idx + 1, len(QUESTIONS))
    text = f"{progress}\n\n{q['q']}"
    if q["t"] == "scale":
        await msg.answer(text, reply_markup=get_scale_keyboard())
    elif q["t"] == "choice":
        await msg.answer(text, reply_markup=get_choice_keyboard(q["opts"]))
    else:
        await msg.answer(text)

# ---------- Команда /export ----------
@router.message(Command("export"))
async def export_excel(message: Message):
    if message.from_user.id != ADMIN_CHAT_ID:
        await message.answer("⛔ У вас нет доступа к этой команде.")
        return

    if not os.path.exists(EXCEL_FILE):
        await message.answer("📭 Таблица пока пуста — никто ещё не прошёл диагностику.")
        return

    try:
        file = FSInputFile(EXCEL_FILE, filename="diagnostics_results.xlsx")
        await message.answer_document(
            file,
            caption=f"📊 *Результаты диагностики*\n🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Ошибка отправки Excel: {e}")
        await message.answer("❌ Не удалось отправить файл. Проверьте логи.")

# ---------- Webhook / Polling ----------
async def set_webhook():
    try:
        await bot.delete_webhook(drop_pending_updates=True)
        await asyncio.sleep(1)
        await bot.set_webhook(WEBHOOK_URL)
        logging.info(f"Вебхук установлен на {WEBHOOK_URL}")
    except Exception as e:
        logging.error(f"Ошибка установки вебхука: {e}")

async def on_startup(bot: Bot):
    if not USE_POLLING:
        await set_webhook()

async def on_shutdown(bot: Bot):
    if not USE_POLLING:
        await bot.delete_webhook()

async def main():
    dp.include_router(router)

    if USE_POLLING:
        logging.info("Запуск в режиме POLLING")
        await bot.delete_webhook(drop_pending_updates=True)
        await dp.start_polling(bot)
    else:
        logging.info("Запуск в режиме WEBHOOK")
        app = web.Application()
        webhook_requests_handler = SimpleRequestHandler(dispatcher=dp, bot=bot)
        webhook_requests_handler.register(app, path=WEBHOOK_PATH)

        app.on_startup.append(lambda _: on_startup(bot))
        app.on_shutdown.append(lambda _: on_shutdown(bot))

        return app

if __name__ == "__main__":
    if USE_POLLING:
        asyncio.run(main())
    else:
        app = asyncio.run(main())
        web.run_app(app, host="0.0.0.0", port=WEBHOOK_PORT)
